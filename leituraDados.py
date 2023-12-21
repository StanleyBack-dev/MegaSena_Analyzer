from openpyxl import load_workbook

def load_data(file_name, sheet_name, start_row, end_row, start_col, end_col):
    # Carregando a planilha
    wb = load_workbook(file_name)

    # Acessando a planilha
    ws = wb[sheet_name]

    # Definindo o intervalo
    data = []
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        data.append([cell.value for cell in row])

    return data

def main():
    file_name = 'MegaSenaGeral.xlsx'
    sheet_name = 'tabelaMegaSena'
    start_row = 8
    end_row = 2675
    start_col = 3
    end_col = 8

    # Lendo os dados do intervalo
    data = load_data(file_name, sheet_name, start_row, end_row, start_col, end_col)

    # Imprimindo os dados
    for row in data:
        print(row)

if __name__ == "__main__":
    main()
